"""
Consolidate Main-Qn-rX groups (Q15..Q20 etc.) across every sheet of an xlsx.
Adds a codeQn column with a data-validation dropdown, and writes a sidecar
.bas file with VBA for multi-select on the code columns.

Usage:
    python formativeTransformation.py

Reads:  input.xlsx        (place next to this script)
Writes: output.xlsx       (open in Excel, Save As -> .xlsm, then paste the .bas)
Writes: output.bas        (paste its contents into ThisWorkbook in the VBA editor)

Skips the sinus-discuss client sheet (metadata; copied verbatim).
"""

import re
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

INPUT  = Path("input.xlsx")
OUTPUT = Path("output.xlsm")   # sidecar .bas has the same stem

JOIN_SEP = "\n"
CODE_SEP = ", "
SKIP_SHEET_PATTERNS = [re.compile(r"sinus[-_ ]?discuss", re.IGNORECASE)]

MAIN_PATTERN = re.compile(
    r"^\s*main\s*[-_ ]?\s*q\s*(\d+)\s*[-_ ]?\s*r\s*(\d+)\b",
    re.IGNORECASE,
)

WRAP = Alignment(wrap_text=True, vertical="top")

# --- Codes ---------------------------------------------------------------
# Add new codes here. To add eSN7 "SomeLabel" and eSN8 "OtherLabel",
# extend the last line's range from range(4, 7) to range(4, 9).
CODES = (
    [f"PBC{i}"  for i in range(1, 14)]       # PBC1..PBC13
  + [f"ePBC{i}" for i in range(14, 17)]      # ePBC14..ePBC16
  + [f"A{i}"    for i in range(1, 5)]        # A1..A4
  + [f"A{i}"    for i in (6, 7, 8, 9)]       # A6..A9 (A5 does not exist)
  + [f"A{i}"    for i in range(51, 53)]      # A51..A52
  + [f"eA{i}"   for i in range(10, 14)]      # eA10..eA13
  + [f"SN{i}"   for i in range(1, 4)]        # SN1..SN3
  + [f"eSN{i}"  for i in range(4, 7)]        # eSN4..eSN6  <-- extend to (4, 9) for eSN7, eSN8
)

CODES_SHEET = "_codes"
CODE_COL_MARK = "_codeColumns"


# --- VBA (paste into ThisWorkbook after saving as .xlsm) -----------------
VBA_SEP = CODE_SEP.replace('"', '""')
VBA_CODE = f'''Option Explicit

Private Sub Workbook_SheetChange(ByVal Sh As Object, ByVal Target As Range)
    On Error GoTo done
    If Target.Cells.Count <> 1 Then Exit Sub
    If Target.Row = 1 Then Exit Sub

    Dim nm As Name
    On Error Resume Next
    Set nm = ThisWorkbook.Names("{CODE_COL_MARK}_" & Sh.Name)
    On Error GoTo done
    If nm Is Nothing Then Exit Sub

    Dim watchedCols As Range
    Set watchedCols = Sh.Range(nm.RefersToRange.Address)
    If Intersect(Target, watchedCols) Is Nothing Then Exit Sub

    Application.EnableEvents = False

    Const SEP As String = "{VBA_SEP}"

    Dim picked As String
    picked = Trim$(CStr(Target.Value))

    Dim existing As String
    If Target.Comment Is Nothing Then
        existing = ""
    Else
        existing = Target.Comment.Text
    End If

    Dim newVal As String
    newVal = ToggleCode(existing, picked, SEP)

    If Not AllTokensValid(newVal, SEP) Then
        MsgBox "Unknown code: '" & picked & "'. Reverted.", vbExclamation, "Invalid code"
        Target.Value = existing
        GoTo cleanup
    End If

    Target.Value = newVal
    If Not Target.Comment Is Nothing Then Target.Comment.Delete
    If Len(newVal) > 0 Then
        Target.AddComment Text:=newVal
        Target.Comment.Visible = False
    End If

cleanup:
    Application.EnableEvents = True
done:
End Sub

Private Function AllTokensValid(val As String, sep As String) As Boolean
    If Len(val) = 0 Then
        AllTokensValid = True
        Exit Function
    End If
    Dim parts() As String, i As Long, t As String, res As Variant
    Dim rng As Range
    Set rng = ThisWorkbook.Sheets("{CODES_SHEET}").Range("A2:A" & (1 + {len(CODES)}))
    parts = Split(val, sep)
    For i = LBound(parts) To UBound(parts)
        t = Trim$(parts(i))
        If Len(t) > 0 Then
            res = Application.Match(t, rng, 0)
            If IsError(res) Then
                AllTokensValid = False
                Exit Function
            End If
        End If
    Next i
    AllTokensValid = True
End Function

Private Function ToggleCode(existing As String, picked As String, sep As String) As String
    If Len(picked) = 0 Then
        ToggleCode = existing
        Exit Function
    End If
    If Len(existing) = 0 Then
        ToggleCode = picked
        Exit Function
    End If
    Dim parts() As String, i As Long, found As Boolean, keep As String
    parts = Split(existing, sep)
    found = False
    keep = ""
    For i = LBound(parts) To UBound(parts)
        If Trim$(parts(i)) = picked Then
            found = True
        Else
            If Len(keep) > 0 Then
                keep = keep & sep & Trim$(parts(i))
            Else
                keep = Trim$(parts(i))
            End If
        End If
    Next i
    If found Then
        ToggleCode = keep
    Else
        ToggleCode = existing & sep & picked
    End If
End Function
'''


# --- Helpers -----------------------------------------------------------------

def should_skip(sheet_name):
    return any(p.search(sheet_name) for p in SKIP_SHEET_PATTERNS)


def classify_headers(headers):
    groups, others = {}, []
    for idx, h in enumerate(headers, start=1):
        if h is None:
            others.append((h, idx))
            continue
        m = MAIN_PATTERN.match(str(h))
        if m:
            q, r = int(m.group(1)), int(m.group(2))
            groups.setdefault(q, []).append((r, idx))
        else:
            others.append((h, idx))
    for q in groups:
        groups[q].sort(key=lambda t: t[0])
        groups[q] = [idx for _, idx in groups[q]]
    return groups, others


def join_cells(ws, row, src_cols):
    parts = []
    for c in src_cols:
        v = ws.cell(row=row, column=c).value
        if v is None:
            continue
        s = str(v).strip()
        if s:
            parts.append(s)
    return JOIN_SEP.join(parts) if parts else None


def copy_sheet_verbatim(ws_in, ws_out):
    for row in ws_in.iter_rows(values_only=False):
        for cell in row:
            ws_out.cell(row=cell.row, column=cell.column, value=cell.value)


def build_codes_sheet(wb):
    ws = wb.create_sheet(title=CODES_SHEET)
    ws["A1"] = "Code"
    for i, code in enumerate(CODES, start=2):
        ws.cell(row=i, column=1, value=code)
    ws.sheet_state = "hidden"
    last_row = 1 + len(CODES)
    return f"={CODES_SHEET}!$A$2:$A${last_row}"


def process_sheet(ws_in, ws_out, log, codes_ref, max_row_hint, wb_out):
    headers = [c.value for c in ws_in[1]]
    groups, others = classify_headers(headers)

    q_list = sorted(groups.keys())
    if not q_list:
        log.append(f"[{ws_in.title}] no Main groups found; copying verbatim")
        copy_sheet_verbatim(ws_in, ws_out)
        return

    matched = ", ".join(f"Q{q}({len(groups[q])})" for q in q_list)
    log.append(f"[{ws_in.title}] Main groups: {matched}")
    log.append(f"[{ws_in.title}] non-Main columns kept: {len(others)}")

    lead, tail = others[:11], others[11:]

    out_headers = [h for h, _ in lead]
    while len(out_headers) < 11:
        out_headers.append(None)
    for q in q_list:
        out_headers.append(headers[groups[q][0] - 1])
        out_headers.append(f"codeQ{q}")
    out_headers += [h for h, _ in tail]

    for col_idx, h in enumerate(out_headers, start=1):
        ws_out.cell(row=1, column=col_idx, value=h)

    main_start = 12
    tail_start = main_start + 2 * len(q_list)

    for r_in in range(2, ws_in.max_row + 1):
        for i, (_, src_idx) in enumerate(lead, start=1):
            ws_out.cell(row=r_in, column=i,
                        value=ws_in.cell(row=r_in, column=src_idx).value)

        for offset, q in enumerate(q_list):
            joined = join_cells(ws_in, r_in, groups[q])
            cell = ws_out.cell(row=r_in,
                               column=main_start + 2 * offset,
                               value=joined)
            cell.alignment = WRAP

        for i, (_, src_idx) in enumerate(tail, start=1):
            ws_out.cell(row=r_in, column=tail_start + i - 1,
                        value=ws_in.cell(row=r_in, column=src_idx).value)

    for offset in range(len(q_list)):
        ws_out.column_dimensions[
            get_column_letter(main_start + 2 * offset)
        ].width = 40
        ws_out.column_dimensions[
            get_column_letter(main_start + 2 * offset + 1)
        ].width = 16

    dv = DataValidation(
        type="list",
        formula1=codes_ref,
        allow_blank=True,
        showDropDown=False,
    )
    dv.prompt = "Pick a code — pick again to remove it."
    dv.promptTitle = "Code"
    ws_out.add_data_validation(dv)

    bottom = max(ws_in.max_row, max_row_hint)
    code_col_letters = []
    for offset in range(len(q_list)):
        col_letter = get_column_letter(main_start + 2 * offset + 1)
        dv.add(f"{col_letter}2:{col_letter}{bottom}")
        code_col_letters.append(f"${col_letter}$2:${col_letter}${bottom}")

    ref = ",".join(f"'{ws_out.title}'!{r}" for r in code_col_letters)
    dn = DefinedName(name=f"{CODE_COL_MARK}_{ws_out.title}", attr_text=ref)
    wb_out.defined_names[dn.name] = dn


def attach_vba(xlsm_path):
    bas_path = xlsm_path.with_suffix(".bas")
    bas_path.write_text(VBA_CODE, encoding="utf-8")
    return bas_path


def main():
    print(f"Reading {INPUT.resolve()}")
    if not INPUT.exists():
        print(f"ERROR: {INPUT} not found.")
        return

    wb_in = load_workbook(INPUT, data_only=True)
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    codes_ref = build_codes_sheet(wb_out)
    max_row_hint = max((wb_in[n].max_row for n in wb_in.sheetnames), default=100) + 500

    log = []
    for name in wb_in.sheetnames:
        ws_out = wb_out.create_sheet(title=name)
        if should_skip(name):
            log.append(f"[{name}] skipped (metadata); copying verbatim")
            copy_sheet_verbatim(wb_in[name], ws_out)
            continue
        process_sheet(wb_in[name], ws_out, log, codes_ref, max_row_hint, wb_out)

    wb_out._sheets = [s for s in wb_out._sheets if s.title != CODES_SHEET] + \
                     [wb_out[CODES_SHEET]]

    xlsx_out = OUTPUT.with_suffix(".xlsx")
    wb_out.save(xlsx_out)

    bas_path = attach_vba(OUTPUT)

    print(f"Wrote {xlsx_out.resolve()}")
    print(f"Wrote VBA module to {bas_path.resolve()}\n")
    print("\n".join(log))
    print()
    print("=" * 60)
    print("ONE-TIME SETUP TO ENABLE MULTI-SELECT")
    print("=" * 60)
    print(f"1. Open  {xlsx_out.name}  in Excel.")
    print(f"2. Save As -> Excel Macro-Enabled Workbook (*.xlsm), keep the name.")
    print(f"3. Press Alt+F11 to open the VBA editor.")
    print(f"4. In the Project pane, right-click 'ThisWorkbook' under your file")
    print(f"   -> View Code. Paste the ENTIRE contents of")
    print(f"      {bas_path.name}")
    print(f"   into that code window (replace anything already there).")
    print(f"5. Close the editor and save. Multi-select is now active on all")
    print(f"   codeQn columns: pick a code to add it, pick it again to remove.")


if __name__ == "__main__":
    main()