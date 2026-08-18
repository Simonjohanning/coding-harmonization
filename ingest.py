"""
Ingest two coded xlsx files (Coder A and Coder B) into versioned JSON.

Usage:
    python ingest.py coderA.xlsx coderB.xlsx [--note "text"] [--replace-all]

Produces:
    public/data/v{N}/answers.json
    public/data/v{N}/codings.json
    public/data/labels.json          (created on v1, shared across versions)
    public/data/log.json             (created on v1, shared across versions)
    public/data/manifest.json        (lists versions + current pointer)

Versioning (default: incremental):
    - Each run creates a new v{N} directory (N = previous + 1).
    - Only COMPLETE respondent/technology pairs present in BOTH coder files are
      considered incoming. One-sided respondents are logged and skipped.
      This allows a round to be ingested even when one coder file also contains
      older respondents, or when one new respondent (e.g. Niki) is not yet coded
      by the other coder.
    - Cells/respondents not part of that matched intersection are copied unchanged
      from the current version.
    - Matched cells are compared with the current version:
        * unchanged coder code sets -> previous harmonization/status carry over
        * changed coder code sets   -> cell is reset to pending
        * completely new cell       -> normal auto/pending handling
    - Discussion entries are always carried forward, tagged with their original
      version.
    - --replace-all restores the old full-snapshot behavior: only cells present
      in the upload are written to the new version.
    - --force-new-version is retained as a deprecated alias for --replace-all.
"""

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

# --- Config ------------------------------------------------------------------

SKIP_SHEET_PATTERNS = [re.compile(r"sinus[-_ ]?discuss", re.IGNORECASE),
                       re.compile(r"^_codes$", re.IGNORECASE)]

MAIN_PATTERN = re.compile(
    r"^\s*main\s*[-_ ]?\s*q\s*(\d+)\s*[-_ ]?\s*r\s*(\d+)\b",
    re.IGNORECASE,
)
# In the output xlsx (from formativeTransformation.py), MainQn headers become
# "Main - QNr1 | ..." (reused from the first rX). codeQn headers are "codeQn".
CODEQ_PATTERN = re.compile(r"^\s*codeQ\s*(\d+)\s*$", re.IGNORECASE)
MAINQ_ANY_PATTERN = re.compile(
    r"main\s*[-_ ]?\s*q\s*(\d+)", re.IGNORECASE
)
# formativeNew.py writes the consolidated question headers simply as
# Q1, Q2, ... Q6 (with codeQ1, codeQ2, ... immediately to the right).
PLAIN_Q_PATTERN = re.compile(r"^\s*Q\s*(\d+)\s*$", re.IGNORECASE)

DATA_DIR = Path("public") / "data"


# --- Helpers -----------------------------------------------------------------

def should_skip(sheet_name):
    return any(p.search(sheet_name) for p in SKIP_SHEET_PATTERNS)


def normalize_tech(sheet_name):
    """Normalize survey sheet titles to the technology ids used online.

    Examples:
        "2 (biomass)"       -> "biomass"
        "3 (fernwärme)"     -> "fernwärme"
        "4 (konventionelle)" -> "konventionelle"
        "6 (wärmepumpe)"    -> "wärmepumpe"

    Plain technology names are left unchanged.
    """
    s = str(sheet_name).strip()
    m = re.match(r"^\s*\d+\s*\((.+)\)\s*$", s)
    return m.group(1).strip() if m else s


def parse_code_cell(v):
    """Comma-separated codes -> sorted unique list (case-preserved)."""
    if v is None:
        return []
    s = str(v).strip()
    if not s:
        return []
    parts = [p.strip() for p in s.split(",")]
    return sorted({p for p in parts if p})


def row_id_from(cells_a_to_k, sheet, row_num):
    """Concatenate column A (username) + column B (externalID) with '|'."""
    a = cells_a_to_k[0]
    b = cells_a_to_k[1]
    a_s = "" if a is None else str(a).strip()
    b_s = "" if b is None else str(b).strip()
    if a_s and b_s:
        return f"{a_s}|{b_s}"
    if a_s or b_s:
        return a_s or b_s
    return f"{sheet}_row{row_num}"


def extract_sheet(ws):
    """Return coded cells for one sheet using a logical question slot (1..6).

    Two transformation formats are supported:
      - old: "Main - Q27r1 | ...", "codeQ27", ...
      - new: "Q1", "codeQ1", ... "Q6", "codeQ6"

    The six recurring questions are matched by their POSITION in the sheet.
    This lets coder files produced by the two transformation scripts be
    compared even though one uses survey question numbers and the other Q1-Q6.
    """
    if should_skip(ws.title):
        return []

    headers = [c.value for c in ws[1]]
    n = len(headers)
    detected = []  # (answer_col, code_col, raw_q_num, is_plain)

    for idx, h in enumerate(headers, start=1):
        if h is None:
            continue
        text = str(h)
        m_main = MAIN_PATTERN.match(text) or MAINQ_ANY_PATTERN.search(text)
        m_plain = PLAIN_Q_PATTERN.match(text)
        m = m_main or m_plain
        if not m:
            continue

        raw_q = int(m.group(1))
        code_idx = None
        for j in range(idx + 1, min(n, idx + 3) + 1):
            hj = headers[j - 1]
            if hj is None:
                continue
            mj = CODEQ_PATTERN.match(str(hj))
            if mj and int(mj.group(1)) == raw_q:
                code_idx = j
                break
        if code_idx is None:
            continue

        detected.append((idx, code_idx, raw_q, bool(m_plain and not m_main)))

    detected.sort(key=lambda x: x[0])
    if not detected:
        return []
    if len(detected) > 6:
        raise ValueError(
            f"[{ws.title}] Found {len(detected)} question/code pairs; expected at most 6."
        )

    tech = normalize_tech(ws.title)
    out = []
    for r in range(2, ws.max_row + 1):
        cells_a_to_k = [ws.cell(row=r, column=c).value for c in range(1, 12)]
        rid = row_id_from(cells_a_to_k, ws.title, r)
        if not any(cells_a_to_k) and not any(
            ws.cell(row=r, column=a_idx).value or ws.cell(row=r, column=c_idx).value
            for a_idx, c_idx, _, _ in detected
        ):
            continue

        for position, (a_idx, c_idx, raw_q, is_plain) in enumerate(detected, start=1):
            # New-format Q1..Q6 explicitly denotes the logical slot. Old-format
            # Main-Qxx columns are also six recurring questions in column order.
            slot = raw_q if is_plain and 1 <= raw_q <= 6 else position
            answer = ws.cell(row=r, column=a_idx).value
            codes = parse_code_cell(ws.cell(row=r, column=c_idx).value)
            out.append({
                "rowId": rid,
                "tech": tech,
                "slot": slot,
                "rawQuestion": f"Q{raw_q}",
                "isPlainQuestion": is_plain,
                "answer": answer if answer is not None else "",
                "codes": codes,
            })
    return out


def load_file(path):
    """Return logical-slot keyed codes/answers plus old-format question mapping."""
    wb = load_workbook(path, data_only=True)
    result = {}
    answers = {}
    question_map = {}  # (tech, slot) -> survey question, e.g. Q27

    for name in wb.sheetnames:
        for entry in extract_sheet(wb[name]):
            key = (entry["rowId"], entry["tech"], entry["slot"])
            result[key] = entry["codes"]
            answers[key] = {
                "tech": entry["tech"],
                "rowId": entry["rowId"],
                "slot": entry["slot"],
                "answer": entry["answer"],
            }
            if not entry["isPlainQuestion"]:
                mk = (entry["tech"], entry["slot"])
                old = question_map.get(mk)
                if old is not None and old != entry["rawQuestion"]:
                    raise ValueError(
                        f"Conflicting question mapping for {entry['tech']} slot Q{entry['slot']}: "
                        f"{old} vs {entry['rawQuestion']}"
                    )
                question_map[mk] = entry["rawQuestion"]
    return result, answers, question_map


def question_num(q):
    m = re.match(r"^Q(\d+)$", str(q).strip(), re.IGNORECASE)
    return int(m.group(1)) if m else None


def previous_slot_maps(prev_codings):
    """Infer the six logical question slots from the current online version.

    Historical biomass data in v5 contains legacy extra cells Q33..Q50. The
    actual recurring six for that technology are the first six (Q27..Q32).
    Therefore, when a technology has more than six historical question ids, we
    keep the six lowest ids for canonical mapping and leave the extra old cells
    untouched in the carried-forward version.
    """
    tech_questions = {}
    if prev_codings:
        for c in prev_codings.get("cells", []):
            n = question_num(c.get("question", ""))
            if n is not None:
                tech_questions.setdefault(c["tech"], set()).add(n)

    q_to_slot = {}
    slot_to_q = {}
    for tech, nums in tech_questions.items():
        ordered = sorted(nums)
        if len(ordered) > 6:
            ignored = ordered[6:]
            print(
                f"NOTE: current {tech} has {len(ordered)} historical question ids; "
                f"using Q{ordered[0]}..Q{ordered[5]} as the six canonical slots "
                f"and carrying {len(ignored)} legacy question ids unchanged."
            )
            ordered = ordered[:6]
        for slot, n in enumerate(ordered, start=1):
            q_to_slot[(tech, f"Q{n}")] = slot
            slot_to_q[(tech, slot)] = f"Q{n}"
    return q_to_slot, slot_to_q


def canonicalize_incoming(codes, answers, question_map, canonical_qmap):
    """Convert logical-slot keyed incoming data to persisted question keys."""
    out_codes = {}
    out_answers = {}
    for (rid, tech, slot), code_list in codes.items():
        q = canonical_qmap.get((tech, slot)) or question_map.get((tech, slot)) or f"Q{slot}"
        key = (rid, tech, q)
        out_codes[key] = code_list
        src = answers[(rid, tech, slot)]
        out_answers[key] = {
            "tech": tech,
            "rowId": rid,
            "question": q,
            "answer": src["answer"],
        }
    return out_codes, out_answers

def cell_id(rowId, tech, question):
    return f"{tech}|{rowId}|{question}"


def key_from_cell(cell):
    return (cell["rowId"], cell["tech"], cell["question"])


def read_json(path):
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2, sort_keys=False),
        encoding="utf-8",
    )


def next_version_number(manifest):
    if not manifest or not manifest.get("versions"):
        return 1
    return max(v["n"] for v in manifest["versions"]) + 1


def carry_discussion(prev, prev_n):
    """Copy previous discussion entries and ensure they have an origin version."""
    discussion = []
    if not prev:
        return discussion
    for d in prev.get("discussion", []):
        if "version" not in d:
            d = {**d, "version": prev_n}
        else:
            d = dict(d)
        discussion.append(d)
    return discussion


# --- Ingest ------------------------------------------------------------------

def ingest(path_a, path_b, note, replace_all=False):
    print(f"Reading {path_a} ...")
    logical_a, logical_answers_a, qmap_a = load_file(path_a)
    print(f"Reading {path_b} ...")
    logical_b, logical_answers_b, qmap_b = load_file(path_b)

    # Previous version must be available locally for an incremental ingest.
    manifest = read_json(DATA_DIR / "manifest.json") or {"current": 0, "versions": []}
    prev_n = manifest.get("current", 0) or 0
    prev_codings = None
    prev_answers = None
    if prev_n:
        prev_codings = read_json(DATA_DIR / f"v{prev_n}" / "codings.json")
        prev_answers = read_json(DATA_DIR / f"v{prev_n}" / "answers.json")
        if prev_codings is None or prev_answers is None:
            raise FileNotFoundError(
                f"Manifest points to v{prev_n}, but public/data/v{prev_n}/answers.json "
                "or codings.json is missing. Refusing incremental ingest."
            )
        print(f"Current local version: v{prev_n}; next version will be v{next_version_number(manifest)}.")
    elif not replace_all:
        print("ERROR: No local current version found (public/data/manifest.json).", file=sys.stderr)
        print("       Incremental ingest would lose the deployed history.", file=sys.stderr)
        print("       Use a project containing the current public/data directory.", file=sys.stderr)
        sys.exit(2)

    prev_q_to_slot, prev_slot_to_q = previous_slot_maps(prev_codings)

    # Resolve a single persisted question ID per technology/slot. Prefer the
    # current online version; otherwise use an old-format incoming workbook.
    canonical_qmap = dict(prev_slot_to_q)
    for source_name, qm in (("Coder A", qmap_a), ("Coder B", qmap_b)):
        for k, q in qm.items():
            if k in canonical_qmap and canonical_qmap[k] != q:
                raise ValueError(
                    f"Question mapping conflict for {k[0]} slot Q{k[1]}: "
                    f"current/other={canonical_qmap[k]}, {source_name}={q}"
                )
            canonical_qmap[k] = q

    codes_a, answers_a = canonicalize_incoming(
        logical_a, logical_answers_a, qmap_a, canonical_qmap
    )
    codes_b, answers_b = canonicalize_incoming(
        logical_b, logical_answers_b, qmap_b, canonical_qmap
    )

    # Match at respondent+technology level, not just individual cells. A row is
    # eligible only when BOTH coder files contain the same complete set of six
    # logical questions. This prevents accidental 5/6 partial respondents.
    def respondent_questions(codes):
        grouped = {}
        for rid, tech, q in codes:
            grouped.setdefault((rid, tech), set()).add(q)
        return grouped

    rq_a = respondent_questions(codes_a)
    rq_b = respondent_questions(codes_b)
    persons_a = set(rq_a)
    persons_b = set(rq_b)
    both_persons = persons_a & persons_b

    matched_persons = set()
    incomplete_or_mismatched = []
    for person in sorted(both_persons):
        qa = rq_a[person]
        qb = rq_b[person]
        if qa == qb and len(qa) == 6:
            matched_persons.add(person)
        else:
            incomplete_or_mismatched.append((person, sorted(qa), sorted(qb)))

    if incomplete_or_mismatched:
        print("WARNING: respondent(s) present in both files but without the same complete six questions; skipped:")
        for (rid, tech), qa, qb in incomplete_or_mismatched[:20]:
            print(f"  {rid} [{tech}] A={qa} B={qb}")

    one_sided_a = sorted(persons_a - persons_b)
    one_sided_b = sorted(persons_b - persons_a)

    incoming_keys = {
        key for key in (set(codes_a) & set(codes_b))
        if (key[0], key[1]) in matched_persons
    }

    print(
        f"Parsed cells: A={len(codes_a)}, B={len(codes_b)}. "
        f"Matched complete respondents={len(matched_persons)} "
        f"({len(incoming_keys)} cells)."
    )
    print(
        f"One-sided respondents skipped: A-only={len(one_sided_a)}, "
        f"B-only={len(one_sided_b)}."
    )
    for label, people in (("A-only", one_sided_a), ("B-only", one_sided_b)):
        for rid, tech in people[:20]:
            print(f"  {label}: {rid} [{tech}]")

    # For matched cells the answer text should also agree. Do not silently merge
    # a respondent if the underlying free-text answer differs between coder files.
    answer_mismatches = []
    for key in sorted(incoming_keys):
        aa = str(answers_a[key].get("answer", "")).strip()
        ab = str(answers_b[key].get("answer", "")).strip()
        if aa != ab:
            answer_mismatches.append((key, aa, ab))
    if answer_mismatches:
        print("ERROR: matched coder rows have different answer text. First mismatches:", file=sys.stderr)
        for key, aa, ab in answer_mismatches[:10]:
            print(f"  {key}: A={aa!r} B={ab!r}", file=sys.stderr)
        raise ValueError("Answer text differs between coder files for matched respondents.")

    prev_index = {}
    prev_keys = set()
    if prev_codings:
        for c in prev_codings.get("cells", []):
            key = key_from_cell(c)
            prev_keys.add(key)
            prev_index[cell_id(*key)] = c

    prev_answers_index = {}
    if prev_answers:
        for a in prev_answers.get("cells", []):
            cid = a.get("cellId") or cell_id(a["rowId"], a["tech"], a["question"])
            prev_answers_index[cid] = a

    # Report whether skipped one-sided respondents are already online or truly
    # new. Existing one-sided rows remain untouched; new ones (such as Niki in
    # round 3) simply wait for the next version when both coders contain them.
    prev_persons = {(rid, tech) for rid, tech, _ in prev_keys}
    skipped_new_a = [p for p in one_sided_a if p not in prev_persons]
    skipped_new_b = [p for p in one_sided_b if p not in prev_persons]
    skipped_existing_a = [p for p in one_sided_a if p in prev_persons]
    skipped_existing_b = [p for p in one_sided_b if p in prev_persons]
    print(
        "Skipped one-sided breakdown: "
        f"existing A-only={len(skipped_existing_a)}, existing B-only={len(skipped_existing_b)}, "
        f"new A-only={len(skipped_new_a)}, new B-only={len(skipped_new_b)}"
    )

    if replace_all:
        all_keys = sorted(incoming_keys)
        print("Mode: full replacement (--replace-all). "
              "Cells missing from the upload will NOT be carried forward.")
    else:
        all_keys = sorted(prev_keys | incoming_keys)
        untouched = prev_keys - incoming_keys
        print(f"Mode: incremental. Carrying {len(untouched)} existing cells "
              "that are absent from the upload.")

    if prev_keys:
        prev_rowids = {k[0] for k in prev_keys}
        incoming_rowids = {k[0] for k in incoming_keys}
        added_rowids = incoming_rowids - prev_rowids
        overlapping_rowids = incoming_rowids & prev_rowids
        print(f"Incoming respondents: {len(incoming_rowids)} "
              f"({len(added_rowids)} new, {len(overlapping_rowids)} already present)")

    new_n = next_version_number(manifest)
    print(f"Creating version v{new_n}...")

    cells = []
    stats = {
        "carried": 0,
        "untouched": 0,
        "reset": 0,
        "auto": 0,
        "pending_new": 0,
        "skipped_one_sided_existing": len(skipped_existing_a) + len(skipped_existing_b),
        "skipped_one_sided_new": len(skipped_new_a) + len(skipped_new_b),
    }

    for key in all_keys:
        rid, tech, q = key
        cid = cell_id(rid, tech, q)
        prev = prev_index.get(cid)
        incoming = key in incoming_keys

        # Incremental mode: if the cell was not part of either new workbook,
        # preserve the current online state verbatim apart from version metadata.
        if prev is not None and not incoming and not replace_all:
            carried = dict(prev)
            carried["discussion"] = carry_discussion(prev, prev_n)
            carried["changedSinceLastVersion"] = False
            carried["carriedFromVersion"] = prev_n
            cells.append(carried)
            stats["carried"] += 1
            stats["untouched"] += 1
            continue

        # Every incoming key belongs to a complete respondent present in both
        # coder files, so both code sets are guaranteed to exist. One-sided
        # respondents never reach this branch.
        ca = codes_a[key]
        cb = codes_b[key]

        discussion = carry_discussion(prev, prev_n)

        # Determine status/harmonization
        codes_unchanged = (
            prev is not None
            and prev.get("codesA", []) == ca
            and prev.get("codesB", []) == cb
        )

        if ca == cb:
            # No discrepancy — auto-resolved
            status = "auto"
            harmonized = list(ca)
            changed = False
            if prev and codes_unchanged and prev.get("status") in {"auto", "resolved"}:
                # Preserve a manually resolved status/harmonization if nothing changed.
                status = prev.get("status", "auto")
                harmonized = prev.get("harmonized", list(ca))
                stats["carried"] += 1
            else:
                stats["auto"] += 1
        elif prev and codes_unchanged:
            # Carry the previous decision (including in-discussion state)
            status = prev.get("status", "pending")
            harmonized = prev.get("harmonized")
            changed = False
            stats["carried"] += 1
        else:
            status = "pending"
            harmonized = None
            changed = prev is not None  # existed before but codes changed
            if changed:
                stats["reset"] += 1
            else:
                stats["pending_new"] += 1

        cells.append({
            "cellId": cid,
            "rowId": rid,
            "tech": tech,
            "question": q,
            "codesA": ca,
            "codesB": cb,
            "harmonized": harmonized,
            "status": status,
            "discussion": discussion,
            "changedSinceLastVersion": changed,
            "carriedFromVersion": prev_n if prev else None,
        })

    # Answers file: incoming answer wins (A before B); cells absent from an
    # incremental upload keep their previous answer text.
    answers_out = []
    for key in all_keys:
        cid = cell_id(*key)
        src = answers_a.get(key) or answers_b.get(key)
        if src:
            answers_out.append({"cellId": cid, **src})
            continue

        prev_answer = prev_answers_index.get(cid)
        if prev_answer is not None and not replace_all:
            answers_out.append(dict(prev_answer))
            continue

        # Defensive fallback: codings should always have a corresponding answer.
        rid, tech, q = key
        answers_out.append({
            "cellId": cid,
            "tech": tech,
            "rowId": rid,
            "question": q,
            "answer": "",
        })

    # Write version files
    vdir = DATA_DIR / f"v{new_n}"
    write_json(vdir / "answers.json", {"version": new_n, "cells": answers_out})
    write_json(vdir / "codings.json", {"version": new_n, "cells": cells})

    # Shared labels + log (only create on v1)
    if not (DATA_DIR / "labels.json").exists():
        write_json(DATA_DIR / "labels.json", {})
    if not (DATA_DIR / "log.json").exists():
        write_json(DATA_DIR / "log.json", {"entries": []})

    # Update manifest
    now = datetime.now(timezone.utc).isoformat()
    manifest["versions"].append({"n": new_n, "created": now, "note": note or ""})
    manifest["current"] = new_n
    write_json(DATA_DIR / "manifest.json", manifest)

    print(f"Wrote {vdir}/")
    print(f"Stats: {stats}")
    print(f"Manifest -> current = v{new_n}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("coder_a", type=Path, help="Coder A xlsx (output of formativeTransformation.py, coded)")
    ap.add_argument("coder_b", type=Path, help="Coder B xlsx")
    ap.add_argument("--note", default="", help="Optional note stored in manifest")
    ap.add_argument(
        "--replace-all",
        "--force-new-version",
        dest="replace_all",
        action="store_true",
        help=(
            "Full-snapshot ingest: only cells in the uploaded files are kept. "
            "--force-new-version is retained as a deprecated alias."
        ),
    )
    args = ap.parse_args()

    if not args.coder_a.exists() or not args.coder_b.exists():
        print("Input file not found", file=sys.stderr)
        sys.exit(1)

    ingest(args.coder_a, args.coder_b, args.note, args.replace_all)


if __name__ == "__main__":
    main()
